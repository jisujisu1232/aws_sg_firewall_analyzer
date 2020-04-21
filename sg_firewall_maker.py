#-*- coding: utf-8 -*-

#Athor : Jisu Kim
'''
python 2.7
pip install boto3 botocore retrying openpyxl requests pyyaml pandas
'''
from __future__ import print_function
import copy
import string
import os
import sys
import openpyxl
import requests
import json
import time
import yaml
import math
import argparse
from datetime import datetime
import pandas as pd

customSeparator = '@#!*%*'
relatedServersStr = 'Related_Servers'
inboundStr = 'ADS_Inbound'
outboundStr = 'ADS_Outbound'
verifiedInbound = 'verified_Inbound'
verifiedOutbound = 'verified_Outbound'
notVerifiedInbound = 'not_' + verifiedInbound
notVerifiedOutbound = 'not_' + verifiedOutbound

def print_elapsed_time(prev_time):
    now=time.time()
    print(now-prev_time)
    return now

def firewall_setter(firewall_infos):
    result = []
    firewall_names = []
    for info in firewall_infos:
        tmp_list = []
        firewall_names.append(info.keys()[0])
        tmp = info[info.keys()[0]]
        for f_idx in range(1,len(tmp.keys())+1):
            for ip_range in tmp[f_idx]:
                tmp_info = []
                tmp_info.append(ip_num_ip_binary(ip_range.split('/')))
                tmp_info.append(f_idx)
                tmp_list.append(tmp_info)
        result.append(tmp_list)
    return [firewall_names,result]

def compare_firewall_and_add(source_ip_binary, target_ip_binary, ads_info_by_hostname, currentColumns, SrcHostname, DestHostname):
    for f_idx, f_info in enumerate(firewall_infos[1]):
        s_info = None
        d_info = None
        for ip_range_info in f_info:
            if s_info == None:
                if source_ip_binary.startswith(ip_range_info[0]):
                    s_info = ip_range_info[1]
            if d_info == None:
                if target_ip_binary.startswith(ip_range_info[0]):
                    d_info = ip_range_info[1]
            if s_info != None and d_info != None:
                if s_info != d_info:
                    if SrcHostname:
                        ads_info_by_hostname[SrcHostname]['firewall_infos']['outbound'][f_idx].append(currentColumns)
                    if DestHostname:
                        ads_info_by_hostname[DestHostname]['firewall_infos']['inbound'][f_idx].append(currentColumns)
                    return
    else:
        pass

def make_related_server(ip, hostname, ip_binary):
    return '{}{}{}{}{}'.format(hostname, customSeparator, ip, customSeparator, ip_binary)


def num_to_binary(num):
     temp = str(bin(num))[2:]
     temp = '0'*(8-len(temp))+temp
     return temp


def ip_num_ip_binary(ip_info):
    prefix = int(ip_info[1])
    ip_classes = ip_info[0].split('.')
    result = ''
    for ip_class in ip_classes:
        result+=num_to_binary(int(ip_class))
    return result[:prefix]


def ip_binary_to_num(ip_binary):
    bLen = len(ip_binary)
    redundancy = 32-bLen
    ip_binary += '0'*redundancy
    result_arr = []
    for i in range(1,5):
        result_arr.append(str(int(ip_binary[8*(i-1):8*i], 2)))
    return '.'.join(result_arr)+'/{}'.format(bLen)


def grouping_ip_range(ip_ranges):
    origin_len = len(ip_ranges)

    ip_ranges.sort()
    final_ip_ranges = []
    final_ip_ranges.append(ip_ranges[0])
    prev_ip_prefix = ip_ranges[0]
    prev_len = len(ip_ranges[0])
    for ip in  ip_ranges[1:]:
        curr_len = len(ip)
        curr_ip_prefix = ip
        if prev_len==curr_len and curr_ip_prefix[:-1]==prev_ip_prefix[:-1]:
            final_ip_ranges.pop()
            curr_ip_prefix = curr_ip_prefix[:-1]
            final_ip_ranges.append(curr_ip_prefix)
        else:
            final_ip_ranges.append(curr_ip_prefix)
        prev_ip_prefix = curr_ip_prefix
        prev_len = curr_len
    if origin_len != len(final_ip_ranges):
        final_ip_ranges=grouping_ip_range(final_ip_ranges)

    return final_ip_ranges


def compare_ip_range(ip32, ip_binary, ip_ranges, customValue):
    for ip_range in ip_ranges:
        if ip_binary.startswith(ip_range):
            return customValue
    return ip32


def check_firewall_application_form_value(form_value):
    # - Allow/Deny : Allow
    # - Expiration_Date : "#add_date# 1Y"
    # - Requester : "#service_admin#"
    # - Remarks : "#hostname# by ADS"
    if form_value:
        form_value = form_value.split(' ')
        for idx, v in enumerate(form_value):
            if len(v)>9 and v.startswith('#') and v.endswith('#'):
                try:
                    dynamic_value_idx = validate_form_values.index(v[1:len(v)-1])
                    if idx==0 and dynamic_value_idx ==len(validate_form_values)-1:
                        addYear = 1
                        if len(form_value)>=2 and form_value[1].isdigit():
                            addYear = int(form_value[1])
                        nowDatetime = datetime.now()
                        nowDatetime = nowDatetime.replace(year=nowDatetime.year+addYear)
                        return [datetime.now().strftime("%Y-%m-%d")]
                except:
                    print("firewall_application_form (config yaml) Dynamic values only support {}.".format(','.join(validate_form_values)))
                    return False
        return form_value
    return ['']


def check_firewall_application_form(firewall_application_form):
    form_keys = []
    form_values = []
    for form in firewall_application_form:
        k = form.keys()[0]
        form_keys.append(k)
        v = check_firewall_application_form_value(form[k])
        if v==False:
            return False
        form_values.append(v)
    return [form_keys, form_values]


def make_firewall_application_form_by_hostname(firewall_application_form, hostname_info):
    form_keys, form_values = copy.deepcopy(firewall_application_form)
    result_values = []
    for value in form_values:
        for idx, v in enumerate(value):
            if len(v)>9 and v.startswith('#') and v.endswith('#'):
                value[idx] = hostname_info[validate_form_values.index(v[1:len(v)-1])]
        result_values.append(' '.join(value))
    return [form_keys, result_values]


def add_related_server(ads_info_by_hostname, hostname, related_server, rs_key, protocol, port):
    if related_server not in ads_info_by_hostname[hostname][relatedServersStr][rs_key].keys():
        ads_info_by_hostname[hostname][relatedServersStr][rs_key][related_server]={}
    if protocol not in ads_info_by_hostname[hostname][relatedServersStr][rs_key][related_server].keys():
        ads_info_by_hostname[hostname][relatedServersStr][rs_key][related_server][protocol]=set()
    ads_info_by_hostname[hostname][relatedServersStr][rs_key][related_server][protocol].add(port)


def port_treat_as_any(output_list, ref_value_of_treat_as_range):
    if len(output_list) >= ref_value_of_treat_as_range and output_list[-1][0].endswith('32'):
        ol = output_list.pop()
        return [[ol[0], ol[1],'Any']]
    return output_list


def create_security_group_by_ads(wb, output_list):
    ws1 = wb.create_sheet()
    global security_group_config
    ip_ranges_treated_equally=security_group_config['ip_ranges_treated_equally']
    ref_value_of_treat_as_range=security_group_config['ref_value_of_treat_as_range']

    ws1.title = 'Security Group'
    ws1.column_dimensions['A'].width = 17
    ws1.column_dimensions['B'].width = 17
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 17
    currentNum = 1
    cellFill = openpyxl.styles.PatternFill(start_color='c0c0c0',
                   end_color='c0c0c0',
                   fill_type='solid')
    cellFont = openpyxl.styles.Font(bold=True)
    cellAlignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    make_col_name_cell(ws1, currentNum, 1, 'Source IP', cellAlignment, cellFill, cellFont)
    make_col_name_cell(ws1, currentNum, 2, 'Protocol', cellAlignment, cellFill, cellFont)
    make_col_name_cell(ws1, currentNum, 3, 'Destination Port', cellAlignment, cellFill, cellFont)

    security_group_output_list = []
    for current in output_list:

        sourceIps = sorted(list(current[0]), key=lambda e: e[1])
        sourceIpLen = len(sourceIps)

        startIdx = 0
        if sourceIpLen >= ref_value_of_treat_as_range:
            sourceBinaryIps = []
            for ip in sourceIps:
                sourceBinaryIps.append(ip[1])
            sourceIps.reverse()
            for ip_prefix in ip_ranges_treated_equally:
                sIdx, eIdx = getStartswithIdx(sourceBinaryIps, ip_prefix[0])
                sourceBinaryIps = sourceBinaryIps[eIdx+1:]
                for i in range(sIdx):
                    security_group_output_list.append(['{}/32'.format(sourceIps.pop()[0]),current[2],current[3]])
                isGroup = (eIdx - sIdx + 1) >= ref_value_of_treat_as_range
                if isGroup:
                    security_group_output_list.append([ip_prefix[1],current[2],current[3]])
                if eIdx > -1:
                    for i in range(eIdx - sIdx + 1):
                        temp = sourceIps.pop()
                        if not isGroup:
                            security_group_output_list.append(['{}/32'.format(temp[0]),current[2],current[3]])
        while sourceIps:
            ip = sourceIps.pop()
            security_group_output_list.append(['{}/32'.format(ip[0]),current[2],current[3]])

    security_group_output_list.sort(key=lambda e: (e[1], e[0]))
    if ref_value_of_treat_as_range:
        port_check_output_list = []
        prevIp = None
        prevProtocol = None
        temp = []
        while security_group_output_list:
            currIp, protocol, port = security_group_output_list.pop()
            if currIp == prevIp and protocol == prevProtocol:
                temp.append([currIp, protocol, port])
            else:
                if temp:
                    port_check_output_list+=port_treat_as_any(temp, ref_value_of_treat_as_range)
                temp=[[currIp, protocol, port]]
            prevIp = currIp
            prevProtocol = protocol
        port_check_output_list+=port_treat_as_any(temp, ref_value_of_treat_as_range)

        security_group_output_list = port_check_output_list
    security_group_output_list.sort(key=lambda e:(e[1],0 if e[2]=='Any' else int(e[2]),e[0]))
    for output in security_group_output_list:
        currentNum+=1
        ip, protocol, port = output
        make_cell(ws1, currentNum, 1, ip).alignment= cellAlignment
        make_cell(ws1, currentNum, 2, 'TCP' if protocol==6 else ('UDP' if protocol==17 else protocol)).alignment= cellAlignment
        make_cell(ws1, currentNum, 3, port).alignment= cellAlignment


def getStartswithIdx(binary_list, prefix):
    cnt = 0
    startIdx=-1
    endIdx=-1
    binaryLen = len(binary_list)
    if binaryLen==0:
        return [-1, -1]
    for idx, binary in enumerate(binary_list):
        if binary.startswith(prefix):
            if startIdx==-1:
                startIdx = idx
            endIdx = idx
        elif binary > prefix:
            return [startIdx, endIdx]
    else:
        if endIdx > -1:
            return [startIdx, endIdx]
        return [len(binary_list), len(binary_list)-1]


sheet_infos=[]
sheet_infos.append(['Security Group', 'AWS security group for To-Be AWS EC2 machine.'])
sheet_infos.append(['FW_Inbound', 'Inbound Information for Firewall Policy Settings.'])
sheet_infos.append(['FW_Outbound', 'Outbound Information for Firewall Policy Settings.'])
sheet_infos.append([verifiedInbound, 'List of servers in the verified IP ranges that send requests to this machine.'])
sheet_infos.append([verifiedOutbound, 'List of servers in the verified IP ranges to which this machine sends requests.'])
sheet_infos.append(['not_{}'.format(verifiedInbound), 'List of servers not in the verified IP ranges that send requests to this machine.'])
sheet_infos.append(['not_{}'.format(verifiedOutbound), 'List of servers not in the verified IP ranges to which this machine sends requests.'])
def create_sheet_info(wb):
    global sheet_infos
    global numberic_verified_ip_range
    global except_ports
    global except_ip_ranges
    ws_info = wb.active
    ws_info.title = 'Sheet Info'
    ws_info.column_dimensions['C'].width = 23
    ws_info.column_dimensions['D'].width = 120
    ws_info.merge_cells(start_row=2, start_column=3, end_row=4, end_column=4)
    ws_info.cell(row=2, column=3, value='Information of the Sheets')
    ws_info.cell(row=2, column=3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws_info.cell(row=2, column=3).font = openpyxl.styles.Font(size=30, bold=True)
    currentCellIdx = 5
    for idx, sheet_info in enumerate(sheet_infos):
        ws_info.cell(row=currentCellIdx, column=3, value='{}. {}'.format(idx+1,sheet_info[0]))
        ws_info.cell(row=currentCellIdx, column=3).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
        ws_info.cell(row=currentCellIdx+1, column=4, value=sheet_info[1])
        ws_info.cell(row=currentCellIdx+1, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
        currentCellIdx+=2
    currentCellIdx+=2
    ws_info.cell(row=currentCellIdx, column=4, value='Verified IP Ranges')
    ws_info.cell(row=currentCellIdx, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    ws_info.cell(row=currentCellIdx, column=4).font = openpyxl.styles.Font(size=15, bold=True)
    currentCellIdx+=1
    if numberic_verified_ip_range:
        for r in numberic_verified_ip_range:
            ws_info.cell(row=currentCellIdx, column=4, value=r)
            ws_info.cell(row=currentCellIdx, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            currentCellIdx+=1
    else:
        ws_info.cell(row=currentCellIdx, column=4, value='None')
        ws_info.cell(row=currentCellIdx, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
        currentCellIdx+=1
    # except_ports
    currentCellIdx+=1
    ws_info.cell(row=currentCellIdx, column=4, value='Exclude Ports')
    ws_info.cell(row=currentCellIdx, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    ws_info.cell(row=currentCellIdx, column=4).font = openpyxl.styles.Font(size=15, bold=True)
    currentCellIdx+=1
    if except_ports:
        for r in except_ports.keys():
            ws_info.cell(row=currentCellIdx, column=4, value=r + ' : ' + ', '.join(str(x) for x in (sorted(except_ports[r]) if except_ports.get(r) else [])))
            ws_info.cell(row=currentCellIdx, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            currentCellIdx+=1
    else:
        ws_info.cell(row=currentCellIdx, column=4, value='None')
        ws_info.cell(row=currentCellIdx, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
        currentCellIdx+=1
    # except_ip_ranges
    currentCellIdx+=1
    ws_info.cell(row=currentCellIdx, column=4, value='Exclude IP Ranges')
    ws_info.cell(row=currentCellIdx, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    ws_info.cell(row=currentCellIdx, column=4).font = openpyxl.styles.Font(size=15, bold=True)
    currentCellIdx+=1
    if except_ip_ranges:
        for r in except_ip_ranges:
            ws_info.cell(row=currentCellIdx, column=4, value=r)
            ws_info.cell(row=currentCellIdx, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            currentCellIdx+=1
    else:
        ws_info.cell(row=currentCellIdx, column=4, value='None')
        ws_info.cell(row=currentCellIdx, column=4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')



def is_In_vpc_range(ip_binary):
    global aws_vpc_ranges
    for r in aws_vpc_ranges:
        if ip_binary.startswith(r):
            return True
    return False

def make_col_name_cell(sheet, cell_row, cell_col, value, alignment, fill, font):
    cell=make_cell(sheet, cell_row, cell_col, value)
    cell.alignment = alignment
    cell.font = font
    cell.fill = fill
    return cell

def make_merge_cell(sheet, start_row, start_column, end_row, end_column, value, alignment):
    sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
    cell = make_cell(sheet, start_row, start_column, value)
    cell.alignment = alignment
    return cell


def make_cell(sheet, cell_row, cell_col, value):
    cell = sheet.cell(row=cell_row, column=cell_col)
    cell.value=value
    return cell

# type
#    0 : inbound
#    1 : outbound
alpha_upper = string.ascii_uppercase
def create_ads_sheet(wb, sheet_name, hostname, ads_info_list, type, firewall_application_form):
    print('\t{} Size : {}'.format(sheet_name, len(ads_info_list)))
    output_list = []
    realType = type
    type = type%2
    host_idx = type+4
    bundleIdx = abs(type-1)
    if ads_info_list:
        final_ads_infos = ads_info_list
        final_ads_infos.sort(key=lambda e: (int(e[2]), int(e[3]), e[-1], e[0]))
        prevProtocol=''
        prevPort=''
        prevIP=''
        temp = ''
        for ads_info in final_ads_infos:
            currentColumns = copy.deepcopy(ads_info)
            currentColumns[type]=set([(currentColumns[type],currentColumns[-2+type],currentColumns[host_idx])])
            tempColumns = temp
            if currentColumns[2]==prevProtocol and currentColumns[3]==prevPort and currentColumns[bundleIdx]==prevIP:
                tempColumns[type].add(next(iter(currentColumns[type])))
                temp = tempColumns
            else:
                output_list.append(temp)
                temp = currentColumns
            prevProtocol=currentColumns[2]
            prevPort=currentColumns[3]
            prevIP=currentColumns[bundleIdx]
        output_list.append(temp)
        output_list = output_list[1:]

    if realType==0:
        create_security_group_by_ads(wb, output_list)
    ws1 = wb.create_sheet()

    ws1.title = sheet_name
    ws1.column_dimensions['A'].width = 17
    ws1.column_dimensions['B'].width = 17
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 17
    currentNum = 1

    cellFill = openpyxl.styles.PatternFill(start_color='c0c0c0',
                   end_color='c0c0c0',
                   fill_type='solid')
    cellAlignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cellFont = openpyxl.styles.Font(bold=True)

    make_col_name_cell(ws1, currentNum, 1, 'Source IP', cellAlignment, cellFill, cellFont)
    make_col_name_cell(ws1, currentNum, 2, 'Destination IP', cellAlignment, cellFill, cellFont)
    make_col_name_cell(ws1, currentNum, 3, 'Protocol', cellAlignment, cellFill, cellFont)
    make_col_name_cell(ws1, currentNum, 4, 'Destination Port', cellAlignment, cellFill, cellFont)
    make_col_name_cell(ws1, currentNum, 5, 'Dest HOSTNAME' if type else 'Src Hostname', cellAlignment, cellFill, cellFont)

    applicationformStartIdx = 5
    for idx, form_key in enumerate(firewall_application_form[0]):
        applicationformStartIdx+=1
        maxlen = max(len(form_key), len(firewall_application_form[1][idx]))
        ws1.column_dimensions[alpha_upper[applicationformStartIdx-1]].width = math.ceil(float(2.07)*maxlen)
        make_col_name_cell(ws1, currentNum, applicationformStartIdx, form_key, cellAlignment, cellFill, cellFont)

    typeColNum=type+1
    bundleColNum=bundleIdx+1
    prevBundleCurrentIP=''
    prev_is_in_vpc=False


    lineNum=0
    output_list.reverse()
    while output_list:
        lineNum+=1
        current = output_list.pop()
        '''
        for lineNum, current in enumerate(output_list, start=1):
        '''
        tempips = list(current[type])
        '''
        if current[bundleIdx]==prevBundleCurrentIP:
            is_in_vpc=prev_is_in_vpc
        else:
            is_in_vpc = is_In_vpc_range(current[-2+bundleIdx])
            prevBundleCurrentIP=current[bundleIdx]
            prev_is_in_vpc=is_in_vpc
        '''
        tempipInfos = []
        print('\t\t{}'.format(lineNum), end='\r')
        for ip in tempips:
            '''
            if '.' in ip[0]:
                ip_binary = ip[1]
                if is_in_vpc and not isVF:
                    if realType > 1 or not is_In_vpc_range(ip_binary):
                        tempipInfos.append(ip)
                else:
                    tempipInfos.append(ip)
            else:
                tempipInfos.append(ip)
        if not tempipInfos:
            continue
            '''
            tempipInfos.append(ip)

        tempipInfos.sort(key=lambda e:e[1])
        ipsLen = len(tempipInfos)
        bundleNum=currentNum+1
        for i, ip_info in enumerate(tempipInfos):
            make_cell(ws1, bundleNum+i, typeColNum, ip_info[0])
            make_cell(ws1, bundleNum+i, 5, ip_info[2])
        mergeNum=currentNum+ipsLen

        if bundleNum==mergeNum:
            make_cell(ws1, bundleNum, bundleColNum, current[bundleIdx]).alignment=cellAlignment
            make_cell(ws1, bundleNum, 3, 'TCP' if current[2]==6 else ('UDP' if current[2]==17 else protocol)).alignment=cellAlignment
            make_cell(ws1, bundleNum, 4, str(current[3])).alignment=cellAlignment
        else:
            make_merge_cell(ws1, bundleNum, bundleColNum, mergeNum, bundleColNum, current[bundleIdx], cellAlignment)
            make_merge_cell(ws1, bundleNum, 3, mergeNum, 3, 'TCP' if current[2]==6 else ('UDP' if current[2]==17 else protocol), cellAlignment)
            make_merge_cell(ws1, bundleNum, 4, mergeNum, 4, str(current[3]), cellAlignment)

        applicationformStartIdx = 5
        for form_value in firewall_application_form[1]:
            applicationformStartIdx+=1
            make_merge_cell(ws1, bundleNum, applicationformStartIdx, mergeNum, applicationformStartIdx, form_value, cellAlignment)
        currentNum += ipsLen


def create_related_servers_sheet(wb, sheet_name, related_servers_list):
    ws2=wb.create_sheet()
    ws2.title = sheet_name
    ws2.column_dimensions['A'].width = 17
    ws2.column_dimensions['B'].width = 17
    ws2.column_dimensions['C'].width = 150
    currentNum = 1
    cellFill = openpyxl.styles.PatternFill(start_color='c0c0c0',
                   end_color='c0c0c0',
                   fill_type='solid')
    cellAlignmentCenter = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cellAlignmentWrap = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
    cellFont = openpyxl.styles.Font(bold=True)
    make_col_name_cell(ws2, currentNum, 1, 'Hostname', cellAlignmentCenter, cellFill, cellFont)
    make_col_name_cell(ws2, currentNum, 2, 'IP', cellAlignmentCenter, cellFill, cellFont)
    make_col_name_cell(ws2, currentNum, 3, 'Ports', cellAlignmentCenter, cellFill, cellFont)
    for related_server in sorted(list(related_servers_list.keys()), key=lambda e: (e.split(customSeparator)[0] if e.split(customSeparator)[0] else 'ZZZ',  e.split(customSeparator)[2])):
        hostname, ip, ip_range = related_server.split(customSeparator)
        currentNum+=1
        make_cell(ws2, currentNum, 1, hostname).alignment = cellAlignmentCenter
        make_cell(ws2, currentNum, 2, ip).alignment = cellAlignmentCenter
        portsStr = ''
        lastProtocolIdx = len(related_servers_list[related_server].keys())-1
        for idx, protocol in enumerate(related_servers_list[related_server].keys()):
            if protocol == 6:
                portsStr+='[TCP]\n'
            elif protocol == 17:
                portsStr+='[UDP]\n'
            else:
                portsStr+=str(protocol)+'\n'
            portsStr+=', '.join(str(x) for x in sorted(list(related_servers_list[related_server][protocol])))
            if lastProtocolIdx > idx:
                portsStr += '\n'
        make_cell(ws2, currentNum, 3, portsStr).alignment = cellAlignmentWrap
        ws2.row_dimensions[currentNum].height = 35*(lastProtocolIdx+1)


risc_widths=[13, 13, 22, 8, 17, 17, 15, 19, 34, 26, 15, 19, 34, 26]
def create_risc_sheet(wb, risc_list):
    ws2=wb.create_sheet()
    ws2.title = 'RISC'
    for i, w in enumerate(risc_widths):
        ws2.column_dimensions[alpha_upper[i]].width = w

    cellFill = openpyxl.styles.PatternFill(start_color='c0c0c0',
                   end_color='c0c0c0',
                   fill_type='solid')
    cellAlignmentCenter = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    cellFont = openpyxl.styles.Font(bold=True)

    column_names = ["Source IP", "Dest IP", "Protocol Name", "Dest Port", "Source Hostname", "Dest Hostname","sour-ce_process","source_application","source_application_context","source_application_instance","dest_process","dest_application","dest_application_context","dest_application_instance", "netstat_count"]
    for colNum, cn in enumerate(column_names, start=1):
        make_col_name_cell(ws2, 1, colNum, cn, cellAlignmentCenter, cellFill, cellFont)

    currentNum = 2
    for risc in risc_list:
        for i, cl in enumerate(risc, start=1):
            make_cell(ws2, currentNum, i, cl).alignment = cellAlignmentCenter
        currentNum+=1

security_group_config = None
numberic_verified_ip_range = None
except_ports = None
except_ip_ranges = None
aws_vpc_ranges = None
validate_form_values=None
isVF = False
firewall_infos = None
# def main(file_name):
def main(arguments):
    global security_group_config
    global numberic_verified_ip_range
    global except_ports
    global except_ip_ranges
    global aws_vpc_ranges
    global validate_form_values
    global isVF
    global firewall_infos
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--type', required=True, choices=['ADS','RISC','RISCALL','VPCFLOW'], help='ADS RISC VPCFLOW')
    parser.add_argument('--source', required=False, help='Full path of connection source .csv file.', default=os.path.join(sys.path[0], 'source.csv'))
    parser.add_argument('--machines', required=False, help='Full path of Machine info .csv file.', default=os.path.join(sys.path[0], 'machine_list.csv'))
    parser.add_argument('--sprint', required=False, help='Migration Sprint')
    args = parser.parse_args(arguments)

    isRISC = (args.type[:4]=='RISC')
    isVF = (args.type=='VPCFLOW')
    if isVF:
        args.sprint = ''
    elif not args.sprint:
        print('[Error] --sprint is required.')
        return
    sprint = str(args.sprint)

    #fileName = os.path.join(sys.path[0], 'machine_list.csv')
    fileName = args.machines
    machine_list = pd.read_csv(fileName, encoding='utf-8', keep_default_na=False)
    validate_form_values = machine_list.columns.tolist()
    validate_form_values.append('add_year')
    yamlPath = os.path.join(sys.path[0], 'sg_firewall_maker.yaml')
    with open(yamlPath, 'r') as ymlfile:
        config = yaml.load(ymlfile, Loader=yaml.FullLoader)
    firewall_application_form = check_firewall_application_form(config['firewall_application_form'] if config.get('firewall_application_form') else [])
    if firewall_application_form==False:
        return

    firewall_infos = firewall_setter(config['firewall_infos']) if config.get('firewall_infos') else []
    isInFirewall = True if firewall_infos else False
    aws_vpc_ranges_temp = config['aws_vpc_ranges'] if config.get('aws_vpc_ranges') else []
    aws_vpc_ranges= []
    for r in aws_vpc_ranges_temp:
        aws_vpc_ranges.append(ip_num_ip_binary(r.split('/')))
    verified_ip_range = []
    numberic_verified_ip_range = config.get('verified_ip_range')
    if numberic_verified_ip_range:
        for ip_range in numberic_verified_ip_range:
            verified_ip_range.append(ip_num_ip_binary(ip_range.split('/')))

    security_group_config = config.get('security_group')
    ref_value_of_treat_as_range=float('inf')
    ip_ranges_treated_equally=[]
    if security_group_config.get('ip_ranges_treated_equally'):
        for ip_range in security_group_config['ip_ranges_treated_equally']:
            ip_ranges_treated_equally.append((ip_num_ip_binary(ip_range.split('/')),ip_range))
        ip_ranges_treated_equally.sort()
        security_group_config['ip_ranges_treated_equally'] = ip_ranges_treated_equally
        security_group_config['ref_value_of_treat_as_range'] = security_group_config['ref_value_of_treat_as_range'] if security_group_config.get('ref_value_of_treat_as_range') else 5
    else:
        security_group_config['ref_value_of_treat_as_range'] = float('inf')

    prev_time=time.time()
    start_time=prev_time
    response = requests.get('https://ip-ranges.amazonaws.com/ip-ranges.json')
    ip_ranges = json.loads(response.text)['prefixes']
    aws_ip_ranges = []
    for ip_range in ip_ranges:
        aws_ip_ranges.append(ip_num_ip_binary(ip_range['ip_prefix'].split('/')))
    final_aws_ip_ranges = grouping_ip_range(aws_ip_ranges)

    resultFolder = os.path.join(sys.path[0], 'results')
    if not os.path.exists(resultFolder):
        os.makedirs(resultFolder)
    resultFolder = os.path.join(resultFolder, datetime.now().strftime("%Y-%m-%d-%H%M%S"))
    if not os.path.exists(resultFolder):
        os.makedirs(resultFolder)
    #athena machine_list
    asis_tobe_ip_dict={}
    all_ip_list={}
    hostname_list = {}
    ads_info_by_hostname = {}

    print("************************")
    print("* Start Machine infos *")
    print("************************")
    for row in  machine_list.values:
        for i, v in enumerate(row):
            if type(v) == unicode:
                row[i] = v.strip()
        hostname = row[0].upper()
        asis_tobe_ip_dict[row[1]]=row[2]
        if isVF or isRISC:
            if row[1]:
                all_ip_list[row[1]]=hostname
            if row[2]:
                all_ip_list[row[2]]=hostname
        if str(row[4]) == sprint or isVF:
            hostname_list[hostname]=row
            ads_info_by_hostname[hostname]={}
            ads_info_by_hostname[hostname][inboundStr]=[]
            ads_info_by_hostname[hostname][outboundStr]=[]
            ads_info_by_hostname[hostname][relatedServersStr]={}
            ads_info_by_hostname[hostname][relatedServersStr][verifiedInbound]={}
            ads_info_by_hostname[hostname][relatedServersStr][verifiedOutbound]={}
            ads_info_by_hostname[hostname][relatedServersStr]['not_{}'.format(verifiedInbound)]={}
            ads_info_by_hostname[hostname][relatedServersStr]['not_{}'.format(verifiedOutbound)]={}
            if isRISC:
                ads_info_by_hostname[hostname]["RISC"]=[]
            if isInFirewall:
                ads_info_by_hostname[hostname]['firewall_infos'] = {}
                ads_info_by_hostname[hostname]['firewall_infos']['inbound'] = []
                ads_info_by_hostname[hostname]['firewall_infos']['outbound'] = []
                for i in range(len(firewall_infos[0])):
                    ads_info_by_hostname[hostname]['firewall_infos']['inbound'].append([])
                    ads_info_by_hostname[hostname]['firewall_infos']['outbound'].append([])
    except_ports = config['exclude_ports'] if config.get('exclude_ports') else {}
    except_ip_ranges = config['exclude_ip_ranges'] if config.get('exclude_ip_ranges') else []
    except_ip_range_binaries = []
    for ir in except_ip_ranges:
        ir = ir.split('/')
        if len(ir)==1:
            ir.append('32')
        except_ip_range_binaries.append(ip_num_ip_binary(ir))

    prev_time=print_elapsed_time(prev_time)
    print("************************")
    print("* End Machine infos *")
    print("************************")

    print("*****************************")
    print("* Start Dependency analysis *")
    print("*****************************")
    #fileName = os.path.join(sys.path[0], args.source)
    fileName = args.source
    if isRISC:
        if len(args.type) == 4:
            f = pd.read_csv(fileName, keep_default_na=False, usecols =["Source IP", "Dest IP", "Protocol Name", "Dest Port", "Source Hostname", "Dest Hostname","source_process","source_application","source_application_context","source_application_instance","dest_process","dest_application","dest_application_context","dest_application_instance"])[["Source IP", "Dest IP", "Protocol Name", "Dest Port", "Source Hostname", "Dest Hostname","source_process","source_application","source_application_context","source_application_instance","dest_process","dest_application","dest_application_context","dest_application_instance"]]
        else:
            f = pd.read_csv(fileName, keep_default_na=False, usecols =["src_addr","dest_addr","protocol_name","dest_port","src_name","dest_name","src_proc","src_app","src_app_context","src_app_instance","dest_proc","dest_app","dest_app_context","dest_app_instance","netstat_count"])[["src_addr","dest_addr","protocol_name","dest_port","src_name","dest_name","src_proc","src_app","src_app_context","src_app_instance","dest_proc","dest_app","dest_app_context","dest_app_instance","netstat_count"]]
    else:
        f = pd.read_csv(fileName, keep_default_na=False)

    except_ports_str = {}
    except_ports_str[6]='TCP'
    except_ports_str[17]='UDP'
    total_cnt = len(f.values)
    next_percent=1

    for l_num, line in enumerate(f.values, start=1):
        if l_num*100/total_cnt==next_percent:
            print('{}%'.format(next_percent), end='\r')
            next_percent+=1
        if isRISC:
            currentColumns = list(line[:6])
        else:
            currentColumns = line
        if isRISC:
            currentColumns[2]=6
            currentColumns[4] = currentColumns[4].upper()
            if currentColumns[4]=='UNKNOWN':
                currentColumns[4]=''
            currentColumns[5] = currentColumns[5].upper()
            if currentColumns[5]=='UNKNOWN':
                currentColumns[5]=''
            source_hostname = all_ip_list.get(currentColumns[0])
            if source_hostname:
                if asis_tobe_ip_dict.get(currentColumns[0]):
                    currentColumns[0] = asis_tobe_ip_dict[currentColumns[0]]
                currentColumns[4] = source_hostname
            destination_hostname = all_ip_list.get(currentColumns[1])
            if destination_hostname:
                if asis_tobe_ip_dict.get(currentColumns[1]):
                    currentColumns[1] = asis_tobe_ip_dict[currentColumns[1]]
                currentColumns[5] = destination_hostname
        elif isVF:
            currentColumns = currentColumns.tolist()
            destination_hostname = all_ip_list.get(currentColumns[1])
            source_hostname = all_ip_list.get(currentColumns[0])
            currentColumns.append(source_hostname if source_hostname else '')
            if source_hostname:
                if asis_tobe_ip_dict.get(currentColumns[0]):
                    currentColumns[0] = asis_tobe_ip_dict[currentColumns[0]]
            if source_hostname:
                if asis_tobe_ip_dict.get(currentColumns[1]):
                    currentColumns[1] = asis_tobe_ip_dict[currentColumns[1]]
            currentColumns.append(destination_hostname if destination_hostname else '')

        source_ip, target_ip, protocol, port, source_hostname, destination_hostname = currentColumns
        if source_ip == target_ip:
            continue
        if protocol in except_ports_str.keys():
            if except_ports.get(except_ports_str[protocol]):
                if int(port) in except_ports[except_ports_str[protocol]]:
                    continue
        source_ip_binary = ip_num_ip_binary([source_ip,'32'])
        target_ip_binary = ip_num_ip_binary([target_ip,'32'])
        for ir in except_ip_range_binaries:
            if source_ip_binary.startswith(ir) or target_ip_binary.startswith(ir):
                break
        else:
            isInSrcHostname = source_hostname and source_hostname in hostname_list.keys()
            isInDestHostname = destination_hostname and destination_hostname in hostname_list.keys()
            currentColumns[1] = compare_ip_range(target_ip, target_ip_binary, final_aws_ip_ranges, 'AWS IP Range')
            currentColumns.append(source_ip_binary)
            currentColumns.append(0 if currentColumns[1] == 'AWS IP Range' else target_ip_binary)
            if isInFirewall and currentColumns[1]!='AWS IP Range':
                compare_firewall_and_add(source_ip_binary, target_ip_binary, ads_info_by_hostname, currentColumns, source_hostname if isInSrcHostname else None, destination_hostname if isInDestHostname else None)

            if isInSrcHostname:
                ads_info_by_hostname[source_hostname][outboundStr].append(currentColumns)
                if currentColumns[1] != 'AWS IP Range':
                    add_related_server(ads_info_by_hostname, source_hostname, make_related_server(target_ip, destination_hostname, target_ip_binary), 'not_{}'.format(verifiedOutbound) if compare_ip_range(target_ip, target_ip_binary, verified_ip_range, None) else verifiedOutbound, protocol, port)
                if isRISC:
                    ads_info_by_hostname[source_hostname]["RISC"].append(line)
            if isInDestHostname:
                ads_info_by_hostname[destination_hostname][inboundStr].append(currentColumns)
                add_related_server(ads_info_by_hostname, destination_hostname, make_related_server(source_ip, source_hostname, source_ip_binary), 'not_{}'.format(verifiedInbound) if compare_ip_range(source_ip, source_ip_binary, verified_ip_range, None) else verifiedInbound, protocol, port)
                if isRISC:
                    ads_info_by_hostname[destination_hostname]["RISC"].append(line)
    prev_time=print_elapsed_time(prev_time)
    print("*****************************")
    print("* End Dependency analysis *")
    print("*****************************")
    all_ip_list.clear()

    print("*******************************")
    print("* Start Excel file generation *")
    print("*******************************")
    no_ads_information_hostnames = []
    for hostname in sorted(hostname_list.keys()):
        print('{}'.format(hostname))
        if len(ads_info_by_hostname[hostname][inboundStr])+len(ads_info_by_hostname[hostname][outboundStr]) == 0:
            no_ads_information_hostnames.append(hostname)
            continue
        firewall_application_form_by_hostname = make_firewall_application_form_by_hostname(firewall_application_form, hostname_list[hostname])
        wb = openpyxl.Workbook()
        tempName = os.path.join(resultFolder, '{}_info.xlsx'.format(hostname))
        create_sheet_info(wb)

        if isInFirewall:
            for i, fn in enumerate(firewall_infos[0]):
                create_ads_sheet(wb, fn+'_FW_Inbound', hostname, ads_info_by_hostname[hostname]['firewall_infos']['inbound'][i], 2, firewall_application_form_by_hostname)
                create_ads_sheet(wb, fn+'_FW_Outbound', hostname, ads_info_by_hostname[hostname]['firewall_infos']['outbound'][i], 3, firewall_application_form_by_hostname)
        create_ads_sheet(wb, 'All_Inbound', hostname, ads_info_by_hostname[hostname][inboundStr], 0, firewall_application_form_by_hostname)

        create_ads_sheet(wb, 'All_Outbound', hostname, ads_info_by_hostname[hostname][outboundStr], 1, firewall_application_form_by_hostname)

        create_related_servers_sheet(wb, verifiedInbound, ads_info_by_hostname[hostname][relatedServersStr][verifiedInbound])

        create_related_servers_sheet(wb, verifiedOutbound, ads_info_by_hostname[hostname][relatedServersStr][verifiedOutbound])

        create_related_servers_sheet(wb, 'not_{}'.format(verifiedInbound), ads_info_by_hostname[hostname][relatedServersStr]['not_{}'.format(verifiedInbound)])

        create_related_servers_sheet(wb, 'not_{}'.format(verifiedOutbound), ads_info_by_hostname[hostname][relatedServersStr]['not_{}'.format(verifiedOutbound)])
        if isRISC:
            create_risc_sheet(wb, ads_info_by_hostname[hostname]['RISC'])
        wb.save(tempName)

        print('{}\n'.format(tempName))
        del ads_info_by_hostname[hostname]

    if no_ads_information_hostnames:
        for h in no_ads_information_hostnames:
            print('No ADS information for hostname[{}]'.format(h))

        tempName = os.path.join(resultFolder, '00.no_information_hostnames.txt')
        f = open(tempName, 'w')
        f.write('\n'.join(no_ads_information_hostnames))
        f.close()
    print_elapsed_time(prev_time)
    print("*******************************")
    print("* End Excel file generation *")
    print("*******************************")
    print("Total : {}".format(time.time()-start_time))



if __name__ == '__main__':
    # main(sys.argv[1:])
    main(sys.argv[1:])
    # ADS
    # main(['--type','ADS','--source','C:\\Users\\jisu7.kim\\Documents\\Project\\ads\\source\\source.csv','--machineInfos','C:\\Users\\jisu7.kim\\Documents\\Project\\ads\\source\\machine_list.csv','--sprint','5'])
    # RISC
    # main(['--type','RISC','--source','C:\\Users\\jisu7.kim\\Documents\\Project\\ads\\source\\source.csv','--machineInfos','C:\\Users\\jisu7.kim\\Documents\\Project\\ads\\source\\machine_list.csv','--sprint','5'])
    # VPCFLOW
    #main(['--type','VPCFLOW','--source','C:\\Users\\jisu7.kim\\Documents\\Project\\ads\\source\\source.csv','--machineInfos','C:\\Users\\jisu7.kim\\Documents\\Project\\ads\\source\\machine_list.csv','--sprint','5'])
    # main(['-h'])
    #python ads.py --source fullpath --machineInfo fullpath
